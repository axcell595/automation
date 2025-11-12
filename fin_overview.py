from __future__ import annotations
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
from datetime import datetime, timedelta
import sys
import logging


# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

def calculate_status(due_date, today):
    """Calculate status based on due date"""
    if pd.isna(due_date):
        return "No Due Date", None
    
    days_diff = (due_date - today).days
    
    if days_diff < 0:
        return "Overdue", "red"
    elif days_diff <= 7:
        return f"Due in {days_diff} days" if days_diff > 0 else "Due Today", "yellow"
    elif days_diff <= 60:
        return f"Due in {days_diff} days", "green"
    else:
        return f"Due in {days_diff} days", None

def process_currency_data(df, currency, today):
    """Process data for a specific currency"""
    # Filter by currency
    currency_df = df[df['Currency'] == currency].copy()
    
    if currency_df.empty:
        return None
    
    # Sort by customer name
    currency_df = currency_df.sort_values('Customer')
    
    # Calculate balance
    currency_df['Balance'] = currency_df['Total'] - currency_df['Paid']
    
    # Add status
    currency_df['Status'], currency_df['Color'] = zip(*currency_df['Due Date'].apply(
        lambda x: calculate_status(x, today)
    ))
    
    return currency_df

def write_to_excel(df, ws, currency, today):
    """Write processed data to Excel worksheet with formatting"""
    # Define styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    
    subtotal_font = Font(bold=True, size=10)
    subtotal_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    grand_total_font = Font(bold=True, size=12)
    grand_total_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    grand_total_font_white = Font(bold=True, size=12, color="FFFFFF")
    
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_font = Font(color="00B050")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['#', 'Customer', 'Invoice #', 'Status', 'Due Date', 'Total', 'Paid', 'Balance']
    ws.append(headers)
    
    # Format header row
    for cell in ws[1]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    row_num = 2
    seq_num = 1
    current_customer = None
    customer_balances = []
    customer_start_row = row_num
    

    for idx, row in df.iterrows():
        
        # Check if we need to add a subtotal
        if current_customer and current_customer != row['Customer']:
            
            # Add subtotal if previous customer had 2+ invoices
            invoice_count = len(customer_balances)
            if invoice_count >= 2:
                ws.append(['', f"Subtotal: {current_customer}", '', '', '', '', '', sum(customer_balances)])
                for col in range(1, 9):
                    cell = ws.cell(row=row_num, column=col)
                    cell.font = subtotal_font
                    cell.fill = subtotal_fill
                    cell.border = border
                    if col == 8:  # Balance column
                        cell.number_format = f'[{currency}] #,##0.00'
                row_num += 1
            
            # Add an empty row after the grouped Customer
            ws.append([''] * 8)
            for col in range (1,9):
                ws.cell(row=row_num, column=col).border = Border()
            row_num += 1
            
            customer_balances = []
            customer_start_row = row_num
        
        current_customer = row['Customer']
        customer_balances.append(row['Balance'])
        
        # Add data row
        data_row = [
            seq_num,
            row['Customer'],
            row['Invoice #'],
            row['Status'],
            row['Due Date'],
            row['Total'],
            row['Paid'],
            row['Balance']
        ]
        ws.append(data_row)
        
        # Format data row
        for col in range(1, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            
            # Number formatting
            if col == 5:  # Due Date
                cell.number_format = 'DD/MM/YYYY'
            elif col in [6, 7, 8]:  # Total, Paid, Balance
                cell.number_format = f'[{currency}] #,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col == 1:  # #
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Apply status color
        status_cell = ws.cell(row=row_num, column=4)
        if row['Color'] == 'red':
            for col in range(1, 9):
                ws.cell(row=row_num, column=col).fill = red_fill
        elif row['Color'] == 'yellow':
            for col in range(1, 9):
                ws.cell(row=row_num, column=col).fill = yellow_fill
        elif row['Color'] == 'green':
            status_cell.font = green_font
        
        row_num += 1
        seq_num += 1
    
    # Add final customer subtotal if needed
    if len(customer_balances) >= 2:
        ws.append(['', f"Subtotal: {current_customer}", '', '', '', '', '', sum(customer_balances)])
        for col in range(1, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.font = subtotal_font
            cell.fill = subtotal_fill
            cell.border = border
            if col == 8:
                cell.number_format = f'[{currency}] #,##0.00'
        row_num += 1
    
    # Add Grand Total
    grand_total = df['Balance'].sum()
    ws.append(['', f"Grand Total Balance: {currency} {grand_total:,.2f}", '', '', '', '', '', grand_total])
    
    ''' This is summary and barchart, revise if needed'''
    summary_start = row_num + 2
    add_customer_balance_chart(ws, df, currency, summary_start)
    
    # Merge cells for grand total label
    ws.merge_cells(f'B{row_num}:G{row_num}')
    
    for col in range(1, 9):
        cell = ws.cell(row=row_num, column=col)
        cell.font = grand_total_font_white if col != 8 else grand_total_font_white
        cell.fill = grand_total_fill
        cell.border = border
        if col == 2:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        if col == 8:
            cell.number_format = f'[{currency}] #,##0.00'
            cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # Adjust column widths
    column_widths = [8, 50, 15, 15, 12, 12, 12, 12]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    logger.info(f'Worksheet for {currency} completed successfully')
    
    
def add_customer_balance_chart(ws, df, currency, start_row):
    '''Writes a Customer vs Balance summary and adds a bar chart.
    start_row: first row to place the summary table.'''
    
    # Build summary
    summary = (
        df.groupby('Customer', as_index=False)['Balance']
          .sum()
          .sort_values('Balance', ascending=False)
    )

    # Header
    ws.cell(row=start_row, column=1, value='Customer').font = Font(bold=True)
    ws.cell(row=start_row, column=2, value='Balance').font = Font(bold=True)

    r = start_row + 1
    for _, rec in summary.iterrows():
        ws.cell(row=r, column=1, value=rec['Customer'])
        bal_cell = ws.cell(row=r, column=2, value=float(rec['Balance']))
        bal_cell.number_format = f'[{currency}] #,##0.00'
        r += 1

    # Create chart
    chart = BarChart()
    chart.title = f'Balance by Customer ({currency})'
    chart.y_axis.title = f'Balance ({currency})'
    chart.x_axis.title = 'Customer'

    # Data range (y values) and category labels (x values)
    data_ref = Reference(ws, min_col=2, min_row=start_row, max_row=r-1)   # Balance col including header
    cats_ref = Reference(ws, min_col=1, min_row=start_row+1, max_row=r-1) # Customers only
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Place the chart (right side of the sheet, adjust anchor as needed)
    ws.add_chart(chart, 'J2')


def generate_report(csv_path, output_path='Financial_Overview_Report.xlsx'):
    """Main function to generate the report"""
    try:
        # Read CSV
        df = pd.read_csv(csv_path)
        
        # Map Xero column names to our expected names
        df.columns = df.columns.str.strip()
        
        # Check if this is a Xero export and map columns
        xero_mapping = {
            'ContactName': 'Customer',
            'InvoiceNumber': 'Invoice #',
            'DueDate': 'Due Date',
            'Total': 'Total',
            'InvoiceAmountPaid': 'Paid',
            'Currency': 'Currency'
        }
        
        # Check if we have Xero columns
        has_xero_columns = all(col in df.columns for col in xero_mapping.keys())
        
        if has_xero_columns:
            logger.info('Detected Xero CSV format - mapping columns...')
            df = df.rename(columns=xero_mapping)
            
            # Group by invoice number and keep only one row per invoice
            # (Xero exports have multiple rows per invoice for line items)
            df = df.groupby('Invoice #').agg({
                'Customer': 'first',
                'Due Date': 'first',
                'Total': 'first',
                'Paid': 'first',
                'Currency': 'first'
            }).reset_index()
            
            logger.info(f'Consolidated to {len(df)} unique invoices')
        else:
            # Standard column names
            required_columns = ['Customer', 'Invoice #', 'Due Date', 'Total', 'Paid', 'Currency']
            missing_cols = [col for col in required_columns if col not in df.columns]
            if missing_cols:
                logger.error(f"Missing required columns: {missing_cols}")
                logger.error(f"Available columns: {list(df.columns)}")
                print(f"Warning: Missing columns: {missing_cols}")
                print(f"Available columns: {list(df.columns)}")
                return
        
        logger.info('All required columns found')
        
        # Parse dates
        df['Due Date'] = pd.to_datetime(df['Due Date'], errors='coerce')
        logger.info('Due dates parsed')
        
        # Convert numeric columns
        df['Total'] = pd.to_numeric(df['Total'], errors='coerce')
        df['Paid'] = pd.to_numeric(df['Paid'], errors='coerce')
        logger.info('Numeric columns converted')
        
        # Fill NaN values
        df['Paid'] = df['Paid'].fillna(0)
        
        # Today's date
        today = datetime(2025, 11, 11)  # As specified
        logger.info(f'Reference date set to: {today.strftime("%Y-%m-%d")}')
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Process USD
        usd_df = process_currency_data(df, 'USD', today)
        if usd_df is not None:
            ws_usd = wb.create_sheet('USD - Payment')
            write_to_excel(usd_df, ws_usd, 'USD', today)
        
        # Process EUR
        eur_df = process_currency_data(df, 'EUR', today)
        if eur_df is not None:
            ws_eur = wb.create_sheet('EUR - Payment')
            write_to_excel(eur_df, ws_eur, 'EUR', today)
        
        # Process other currencies if any
        other_currencies = df[~df['Currency'].isin(['USD', 'EUR'])]['Currency'].unique()
        if len(other_currencies) > 0:
            for curr in other_currencies:
                curr_df = process_currency_data(df, curr, today)
                if curr_df is not None:
                    ws_other = wb.create_sheet(f'{curr} - Payment')
                    write_to_excel(curr_df, ws_other, curr, today)
        
        # Save workbook
        wb.save(output_path)
        print(f"✓ Report generated successfully: {output_path}")
        
    except Exception as e:
        print(f"Error generating report: {str(e)}")
        import traceback
        traceback.print_exc()

# Usage
if __name__ == "__main__":
    if len(sys.argv) > 1:
        csv_file = sys.argv[1]
    else:
        csv_file = "xero.csv"  # Default filename
    
    generate_report(csv_file)
