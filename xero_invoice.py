"""
Invoice Processing Script - Improved Version
Processes Xero invoice exports and generates formatted Excel reports
"""

import pandas as pd
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment
import shutil
import logging
from typing import Tuple

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class InvoiceProcessor:
    """Handles invoice data processing and Excel generation."""
    
    # Configuration constants
    INPUT_FILE = 'invoices.csv'
    TEMPLATE_FILE = 'template.xlsx'
    OUTPUT_FILE = 'fo_out.xlsx'
    
    COLUMN_MAPPING = {
        'ContactName': 'Customer Name',
        'InvoiceNumber': 'Invoice #',
        'Status': 'Payment Status',
        'DueDate': 'Due date',
        'Total': 'Total Amount',
        'InvoiceAmountPaid': 'Amount Paid',
        'InvoiceAmountDue': 'Balance',
        'Description': 'Description'
    }
    
    FINAL_COLUMNS = ['#', 'Customer Name', 'Invoice #', 'Payment Status', 
                     'Due date', 'Total Amount', 'Amount Paid', 'Balance', 'Description']
    
    CURRENCY_FORMATS = {
        'EUR': '"€" #,##0',
        'USD': '"$" #,##0'
    }
    
    def __init__(self):
        self.today = datetime.now().date()
        
    def load_and_clean_data(self) -> pd.DataFrame:
        """Load CSV data and perform initial cleaning."""
        try:
            df = pd.read_csv(self.INPUT_FILE, encoding='utf-8')
            logger.info(f"Loaded {len(df)} records from {self.INPUT_FILE}")
        except FileNotFoundError:
            logger.error(f"Input file '{self.INPUT_FILE}' not found")
            raise
        except Exception as e:
            logger.error(f"Error loading CSV: {e}")
            raise
        
        # Clean customer names
        df['ContactName'] = df['ContactName'].replace(
            "Integral Poland SpÃƒÂ³Ã…â€ška z ograniczonÃ„â€¦ odpowiedzialnoÃ…â€ºciÃ„â€¦",
            "Integral Poland"
        )
        
        # Remove duplicates
        initial_count = len(df)
        df = df.drop_duplicates(subset=['ContactName', 'InvoiceNumber'])
        removed = initial_count - len(df)
        if removed > 0:
            logger.info(f"Removed {removed} duplicate records")
        
        return df
    
    def filter_outstanding_invoices(self, df: pd.DataFrame) -> pd.DataFrame:
        """Filter for outstanding invoices only."""
        # Convert DueDate with proper format handling
        df['DueDate'] = pd.to_datetime(df['DueDate'], dayfirst=True, errors='coerce')
        
        # Filter outstanding invoices
        mask = (df['InvoiceAmountDue'] > 0) & (df['Status'] == 'Awaiting Payment')
        filtered_df = df[mask].copy()
        
        logger.info(f"Filtered to {len(filtered_df)} outstanding invoices")
        return filtered_df
    
    @staticmethod
    def calculate_status(due_date: pd.Timestamp, today: datetime.date) -> str:
        """Calculate payment status based on due date."""
        if pd.isna(due_date):
            return ''
        
        delta = (due_date.date() - today).days
        
        if delta == 0:
            return "Due today"
        elif delta > 0:
            day_str = "day" if delta == 1 else "days"
            return f"Due in {delta} {day_str}"
        else:
            abs_delta = abs(delta)
            day_str = "day" if abs_delta == 1 else "days"
            return f"Overdue by {abs_delta} {day_str}"
    
    def prepare_data(self, df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Prepare and split data into EUR and USD dataframes."""
        # Calculate status
        df['Status'] = df['DueDate'].apply(
            lambda d: self.calculate_status(d, self.today)
        )
        
        # Select and order columns
        required_cols = list(self.COLUMN_MAPPING.keys()) + ['Currency']
        df = df[required_cols].copy()
        
        # Split by currency
        eur_df = df[df['Currency'] == 'EUR'].drop('Currency', axis=1)
        usd_df = df[df['Currency'] == 'USD'].drop('Currency', axis=1)
        
        logger.info(f"Split data: {len(eur_df)} EUR invoices, {len(usd_df)} USD invoices")
        
        # Process each currency dataframe
        eur_df = self._format_dataframe(eur_df)
        usd_df = self._format_dataframe(usd_df)
        
        return eur_df, usd_df
    
    def _format_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Format a single currency dataframe."""
        if df.empty:
            # Return empty dataframe with correct columns
            empty_df = pd.DataFrame(columns=self.FINAL_COLUMNS)
            return empty_df
        
        # Rename columns
        df = df.rename(columns=self.COLUMN_MAPPING)
        
        # Format dates
        df['Due date'] = df['Due date'].apply(
            lambda d: d.strftime('%d-%m-%Y') if pd.notna(d) else ''
        )
        
        # Sort by due date
        df = df.sort_values(by='Due date', ascending=True).reset_index(drop=True)
        
        # Add row numbers
        df.insert(0, '#', range(1, len(df) + 1))
        
        return df
    
    def update_excel(self, eur_df: pd.DataFrame, usd_df: pd.DataFrame):
        """Update Excel workbook with processed data."""
        # Copy template
        try:
            shutil.copy(self.TEMPLATE_FILE, self.OUTPUT_FILE)
            logger.info(f"Created output file from template")
        except FileNotFoundError:
            logger.error(f"Template file '{self.TEMPLATE_FILE}' not found")
            raise
        
        # Load workbook
        wb = openpyxl.load_workbook(self.OUTPUT_FILE)
        
        # Update sheets
        self._update_sheet(wb['EUR'], eur_df, 'EURTable', 'EUR_Pivot', 'EUR')
        self._update_sheet(wb['USD'], usd_df, 'USDTable', 'USD_Pivot', 'USD')
        
        # Save
        wb.save(self.OUTPUT_FILE)
        logger.info(f"Successfully saved output to {self.OUTPUT_FILE}")
    
    def _update_sheet(self, ws, df: pd.DataFrame, table_name: str, 
                      pivot_sheet_name: str, currency: str):
        """Update a single worksheet with data."""
        # Clear existing data (rows 4+)
        if ws.max_row > 3:
            ws.delete_rows(4, ws.max_row - 3)
        
        num_rows = len(df)
        last_data_row = 3 + num_rows
        
        # Write data
        for row_idx, row_data in enumerate(df.itertuples(index=False), start=4):
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Add blank row
        ws.insert_rows(last_data_row + 1)
        blank_row = last_data_row + 1
        
        # Copy formatting from row 2 to blank row
        self._copy_row_format(ws, source_row=2, target_row=blank_row)
        
        # Add total row
        total_row = last_data_row + 2
        self._add_total_row(ws, total_row, last_data_row, currency)
        
        # Clean up rows below total
        if ws.max_row > total_row:
            ws.delete_rows(total_row + 1, ws.max_row - total_row)
        
        # Copy formatting for extended rows (from template row 80)
        for offset in range(1, 22):
            self._copy_row_format(ws, source_row=80, target_row=total_row + offset)
        
        # Format data rows
        if num_rows > 0:
            for row in range(4, total_row):
                self._copy_row_format(ws, source_row=4, target_row=row)
        
        # Format specific cells
        source_style = ws.cell(2, 1)
        target_cell = ws.cell(total_row, 9)
        if source_style.has_style:
            target_cell._style = source_style._style
        
        # Update table
        self._update_table(ws, table_name, last_data_row, num_rows)
        
        # Set pivot refresh
        self._set_pivot_refresh(ws.parent, pivot_sheet_name)
    
    def _copy_row_format(self, ws, source_row: int, target_row: int):
        """Copy formatting from source row to target row."""
        for col in range(1, ws.max_column + 1):
            source_cell = ws.cell(source_row, col)
            target_cell = ws.cell(target_row, col)
            if source_cell.has_style:
                target_cell._style = source_cell._style
    
    def _add_total_row(self, ws, total_row: int, last_data_row: int, currency: str):
        """Add total calculation row."""
        label = "GRAND TOTAL" if currency == 'EUR' else f"TOTAL {currency}"
        ws.cell(total_row, 7, value=label)
        
        # Set font
        font = Font(name="Aptos", bold=True, size=14)
        ws.cell(total_row, 7).font = font
        
        # Add sum formula (column H = 8)
        sum_range = f"H4:H{last_data_row}"
        ws.cell(total_row, 8).value = f"=SUBTOTAL(9,{sum_range})"
        ws.cell(total_row, 8).number_format = self.CURRENCY_FORMATS[currency]
        ws.cell(total_row, 8).font = font
        ws.cell(total_row, 8).alignment = Alignment(horizontal='right')
    
    def _update_table(self, ws, table_name: str, last_row: int, num_rows: int):
        """Update or create Excel table."""
        table_ref = f"A3:I{last_row}" if num_rows > 0 else "A3:I3"
        
        if table_name in ws.tables:
            ws.tables[table_name].ref = table_ref
        else:
            tab = Table(displayName=table_name, ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tab.tableStyleInfo = style
            ws.add_table(tab)
    
    def _set_pivot_refresh(self, wb, pivot_sheet_name: str):
        """Enable pivot table refresh on workbook open."""
        if pivot_sheet_name in wb.sheetnames:
            pivot_ws = wb[pivot_sheet_name]
            for pivot in pivot_ws._pivots:
                pivot.cache.refreshOnLoad = True
    
    def process(self):
        """Main processing pipeline."""
        try:
            logger.info("Starting invoice processing...")
            
            # Load and process data
            df = self.load_and_clean_data()
            df = self.filter_outstanding_invoices(df)
            eur_df, usd_df = self.prepare_data(df)
            
            # Update Excel
            self.update_excel(eur_df, usd_df)
            
            logger.info("Processing complete!")
            print(f"\n✓ Processing complete! Output saved to {self.OUTPUT_FILE}")
            print(f"  EUR invoices: {len(eur_df)}")
            print(f"  USD invoices: {len(usd_df)}")
            print(f"  Open in Excel to see updated pivots.")
            
        except Exception as e:
            logger.error(f"Processing failed: {e}")
            raise


def main():
    """Entry point for the script."""
    processor = InvoiceProcessor()
    processor.process()


if __name__ == '__main__':
    main()
